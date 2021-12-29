# Iago Leonardo Alves de Oliveira
# 04/04/2021

# Bot desenvolvido para buscar infomações no mercado livre e anotar no excel
# Há duas opções: 
# + Criar um novo catálogo de vendedor no excel
# + Atualizar um catálogo já existente, também no excel
# 
# A opção de mapear requer o caminho onde salvar o documento excel, o link e nome do vendedor na qual deseja mapear
# Já a opção de atualizar, requer apenas o caminho da planilha na qual deseja atualizar
# (Ambas são executadas a partir de bibliotecas criadas especificamente para estas tarefas)

#### Importa as bibliotecas
#OS - Ações do sistema operacional (ja vem com python)
import os, sys
#Selenium - Ações no chrome (pip install selenium)
from selenium import webdriver
#WebDriverManager - Instalar chrome driver automatico (pip install webdriver_manager)
from webdriver_manager.chrome import ChromeDriverManager
#PyQt5 - Interface gráfica (.ui)
from PyQt5 import uic, QtWidgets
#OpenPyXl - Ações no excel
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
#tkinker - Abrir explorador de arquivos
from tkinter import Tk
from tkinter.filedialog import askdirectory, askopenfilename
#PyAutoGui simular ações de usuário (usei para criar caixas de dialogo)
import pyautogui
#Manipulação de Datas
from datetime import date


#Bibliotecas exclusivas
from Libs.catalogo import buscarCatalogo
from Libs.vendasEstoque import buscarVendasEstoque

#Recebe a data atual
data = date.today()
#Formata (dd/mm/aaaa)
data = data.strftime('%d/%m/%Y')

#Cria a cor de fundo
azul = PatternFill(start_color='0D1385', end_color='0D1385', fill_type='solid')

#Tenta abrir o navegador
try:
    #Implementa configurações
    options = webdriver.ChromeOptions()
    options.add_argument('lang=pt-br') #Portugues
    options.add_argument('--disable-notifications') #Sem notificações
    #Aplica opções que removem o aviso 'o chrome está sendo controle por um software'
    options.add_experimental_option('excludeSwitches', ['enable-automation'])
    options.add_experimental_option('useAutomationExtension', False)

    #Baixa o chrome driver de acordo com a versão do chrome do usuario
    navegador = webdriver.Chrome(ChromeDriverManager().install(), options=options)

    os.system('cls') #Limpa o terminal
    
    #Maximiza a tela
    navegador.maximize_window()

#Se der erro
except Exception as erro:
    #Exibe mensagem de erro
    pyautogui.alert(text='Algo deu errado com o ChromeDriver\n{}'.format(erro), title='ERRO!', button='OK')

    
#Função que exibe o input de ID VENDEDOR
def exibirInput():
    #Exibe os valores
    forms.editID.setVisible(True)
    forms.editNome.setVisible(True)
    forms.txtID.setText('Link do Perfil do Vendedor')
    forms.txtNome.setText('Nome do Vendedor')

    #Altera o caminho na exibição
    forms.txtCaminho.setText('Caminho onde salvar a planilha')

    #Altera o texto do botao
    forms.btnExecutar.setText('Mapear')

#Função que oculta o input de ID VENDEDOR
def ocultarInput():
    #Oculta os inputs
    forms.editID.setVisible(False)
    forms.editNome.setVisible(False)
    forms.txtID.setText('')
    forms.txtNome.setText('')

    #Altera o caminho na exibição
    forms.txtCaminho.setText('Caminho onde salvar a planilha')

    #Altera o texto do botao
    forms.btnExecutar.setText('Atualizar')

#Função que recebe o caminho onde salvar o arquivo 
def alterarCaminho():
    # oculta a janela tkinter
    Tk().withdraw() 

    #Se ele for mapear um catalogo
    if forms.rd_mapear.isChecked():
        #Tenta receber a pasta
        try:    
            #Recebe a pasta selecionada
            caminho = askdirectory(title="Selecionar pasta")

            #Se tiver selecionado uma pasta
            if caminho != '':
                #Altera o caminho na exibição
                forms.txtCaminho.setText(caminho)
        
        #Caso dê erro
        except Exception as erro:
            #Exibe mensagem de erro
            pyautogui.alert(text='Erro ao receber pasta\n({})'.format(erro), title='ERRO!', button='OK')
            #Fecha o navegador
            navegador.close()
            #Encerra
            sys.exit(0)

    #Se ele for atualizar vendas
    elif forms.rd_atualizar.isChecked():
        #Tenta receber o caminho da planilha
        try:
            #Abre o gerenciador de arquivos para o usuario selecionar a planilha
            caminho = askopenfilename(title = "Selecione a planilha que deseja atualizar" , filetypes = (( "Arquivos Excel" , ".xlsx" ), ( "todos os arquivos" , "* . * " ))) # selecionar um arquivo

            #Se tiver selecionado uma pasta
            if caminho != '':
                #Altera o caminho na exibição
                forms.txtCaminho.setText(caminho)

        #Caso dê erro, exibe ao usuario
        except Exception as erro:
            #Exibe mensagem de erro
            pyautogui.alert(text='Erro ao receber arquivo\n{}'.format(erro), title='ERRO!', button='OK')


#Função principal
def executar():

    #Recebe o caminho do arquivo
    caminho = forms.txtCaminho.text()

    #Verifica se selecionou um caminho
    if caminho == 'Caminho onde salvar a planilha':
        #Exibe mensagem de erro
        pyautogui.alert(title='Atenção!', text='Informe o caminho do arquivo antes de continuar', button='OK')

    
    #Senão, continua
    else:
        ##########################################
        ####### ATUALIZAR VENDAS E ESTOQUE #######
        ########################################## 
        if forms.rd_atualizar.isChecked() :
            
            #Variaveis que serão usada no excel
            pastaExcel = ''
            planilha = ''

            #Tenta carregar o arquivo excel
            try:
                #Carrega a pasta do Excel
                pastaExcel = load_workbook(caminho)
            
                #Recebe o nome da guia do excel
                guia = pastaExcel.sheetnames[0]
                #Definindo a planilha da Pasta de trabalho a partir do nome recebido
                planilha = pastaExcel[guia]

                #planAberta recebe true
                planAberta = True
                
            #Caso dê erro, exibe ao usuario
            except Exception as erro:
                #Exibe mensagem de erro
                pyautogui.alert(title='ERRO!', text='Erro ao abrir arquivo Excel \n (Talvez esteja aberto)\n{}'.format(erro), button='OK')
                #planAberta recebe false para a execução nao continuar
                planAberta = False

            if planAberta:
                #Contador se inicia em 1
                cont = 1
                #Especifica a linha inicial e as colunas
                linha = 2
                #Especifica as colunas
                colProduto = 1
                colValor = 2
                colEstoque = 3
                colLink = 4
                #Coluna data será definida durante a execução
                colData = 1

                #Procura uma coluna vazia na linha 1
                while planilha.cell(row=1, column=colData).value != None:
                    #Avança 1 coluna
                    colData += 1

                #Insere a data na coluna onde não irá sobrescrever (na primeira coluna vazia que encontrou)
                planilha.cell(row=1, column=colData, value=data)
                #Cor de fundo 
                planilha.cell(row=1, column=colData).fill = azul
                #Fonte (tamanho 11 / Branca / Negrito)
                planilha.cell(row=1, column=colData).font = Font(size=11, color='FFFFFF', bold=True)
                #Alinhamento de Texto (centralizado)
                planilha.cell(row=1, column=colData).alignment = Alignment(horizontal='center', vertical='center')

            
                #Enquanto a celula do link não estiver vazia
                while planilha.cell(row=linha, column=colLink).value != None:
                    
                    #Recebe o link em formato de texto
                    linkProduto = str(planilha.cell(row=linha, column=colLink).value)
                    
                    #Recebe a quantidade de vendas do anuncio do link
                    vendas_estoque = buscarVendasEstoque(linkProduto, navegador)
                    
                    #Anota o valor no excel
                    planilha.cell(row=linha, column=colData, value=vendas_estoque[0])
                    planilha.cell(row=linha, column=colEstoque, value=vendas_estoque[1])

                    #Desce uma linha
                    linha += 1
                    
                    #Recebe as informações do produto
                    nomeProduto = str(planilha.cell(row=linha, column=colProduto).value)
                    valorProduto = str(planilha.cell(row=linha, column=colValor).value)
                    
                    #Exibe no terminal
                    print('\n\t\t\tEste é o item {}'.format(str(cont)))
                    print('\n Produto: {} \n Valor: {} \n Estoque: {} \n LINK: {} \n Vendas: {} \n'.format(nomeProduto, valorProduto, vendas_estoque[1], linkProduto, vendas_estoque[0]))
                    
                    #Contador recebe mais 1
                    cont += 1

                #Tenta salvar o excel
                try:
                    pastaExcel.save(caminho)
                    pastaExcel.close()
                #Se der erro, pede para fechar a planilha
                except:
                    #Exibe mensagem
                    pyautogui.alert(title='ATENÇÃO!', text='O bot não está conseguindo salvar as alterações\nFecha a planilha e clique em ok após ela estar fechada', button='Já fechei a planilha')
                    #Tenta salvar o excel de novo
                    try:
                        pastaExcel.save(caminho)
                        pastaExcel.close()
                    #Se der erro, informa
                    except:
                        #Exibe mensagem de erro
                        pyautogui.alert(title='ERRO!', text='Não foi possível salvar as alterações\nDetalhes do erro: {}'.format(erro), button='OK')
                


        ##########################################
        ########    MAPEAR CATALOGO      #########
        ########################################## 
        elif forms.rd_mapear.isChecked() :
            
            #Recebe o nome e ID do vendedor
            nomeVendedor = forms.editNome.text()
            linkVendedor = forms.editID.text()

            #Formata os valores, remove caracteres especiais
            linkVendedor = linkVendedor.replace(' ', '')
            linkVendedor = linkVendedor.replace('\n', '')
            nomeVendedor = nomeVendedor.replace(' ', '_')
            nomeVendedor = nomeVendedor.replace('/', '_')
            nomeVendedor = nomeVendedor.replace('\\', '_')
            nomeVendedor = nomeVendedor.replace('$', '_')
            nomeVendedor = nomeVendedor.replace('%', '_')
            nomeVendedor = nomeVendedor.replace('#', '_')
            nomeVendedor = nomeVendedor.replace('@', '_')
            nomeVendedor = nomeVendedor.replace('!', '_')
            nomeVendedor = nomeVendedor.replace('&', '_')
            nomeVendedor = nomeVendedor.replace(',', '_')
            nomeVendedor = nomeVendedor.replace('*', '_')
            nomeVendedor = nomeVendedor.replace('|', '_')
            nomeVendedor = nomeVendedor.replace(';', '_')
            nomeVendedor = nomeVendedor.replace('.', '_')
            nomeVendedor = nomeVendedor.replace('\n', '')

            #Verifica se um id e nome de vendedor foi inserido
            if linkVendedor == '' or nomeVendedor == '':
                #Exibe mensagem
                pyautogui.alert(text='Insira Link e Nome do Vendedor', title='Alerta!', button='OK')
            
            else:
                #Chama a função que busca as infos do vendedor e anota no excel
                buscarCatalogo(linkVendedor, nomeVendedor, navegador, caminho)

            #Limpa os edit texts
            forms.editNome.setText('')
            forms.editID.setText('')



        #Altera o caminho na exibição
        forms.txtCaminho.setText('Caminho onde salvar a planilha')
        #Informa que a execução chegou ao fim
        pyautogui.alert(title='Fim da Execução!', text='A execução chegou ao fim!', button='OK')



###################################
############ EXECUÇÃO #############
###################################

app = QtWidgets.QApplication([])
#Carrega a interface gráfica
forms = uic.loadUi(uifile=os.path.dirname(os.path.abspath(__file__)) + '/interface/forms.ui')
#Quando o botão for pressionado, chama a função
forms.btnExecutar.clicked.connect(executar)
forms.btnAlterar.clicked.connect(alterarCaminho)
#Ao clicar nos radiobuttons ativa/desativa os inputs de nome e id 
forms.rd_mapear.toggled.connect(exibirInput)
forms.rd_atualizar.toggled.connect(ocultarInput)

#Exibe a interface
forms.show()
#Executa
app.exec()