# Iago Leonardo Alves de Oliveira
# 04/04/2021

# Biblioteca desenvolvida para buscar infomações de anuncios de um vendedor no mercado livre
# As informações são anotadas em uma planilha do excel

# Passo a passo:
# + Recebe o caminho da pasta onde a planilha será salva
# + Cria a planilha do excel com uma aba nomeada 'Catálogo'
# + Define os cabeçalhos da planilha onde as informações serão inseridas (nome das colunas)
# + Acessa o perfil do vendedor através do link, utilizando o chrome driver
# + Realiza um loop, acessando cada item do catalogo do vendedor
# + É extraído: Nome do Produto; Estoque disponível; Valor; Link; Número de vendas na data da execução
# + Quando chegar ao ultimo item da pagina, o bot muda para a próxima página (cada pagina tem 48 itens)
# + Caso o mercado livre derrube-o devido a muitas solicitações, o bot para por um periodo e depois volta a ativa 
# + Ao finalizar, a planilha é salva no caminho especificado

####Importando as bibliotecas
#OpenPyXl - Ações no excel
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
#PyAutoGui simular ações de usuário (usei para criar caixas de dialogo)
import pyautogui
#Manipulação de Datas
from datetime import date

#Biblioteca exclusiva
from Libs.limparNavegador import burlarBarreira


#Recebe a data atual
data = date.today()
#Formata (dd/mm/aaaa)
data = data.strftime('%d/%m/%Y')

#Cria a cor de fundo
azul = PatternFill(start_color='0D1385', end_color='0D1385', fill_type='solid')


#Função que busca informações dos anuncios do vendedor e anota no excel
def buscarCatalogo(linkVendedor, nomeVendedor, navegador, caminho):

    #=============================
    #======== CRIA EXCEL =========
    #=============================

    #Especifica a linha inicial
    linha = 2
    #Especifica as colunas
    colProduto = 1
    colValor = 2
    colEstoque = 3
    colLink = 4
    colData = 5

    #Cria a pasta de trabalho
    pastaExcel = Workbook()
    
    #Define a primeira planilha
    planilha = pastaExcel.active
    #Define o nome da planilha
    planilha.title = 'Catalogo'
    
    #Define os cabeçalhos
    planilha.cell(row=1, column=colProduto, value='Produto')
    planilha.cell(row=1, column=colValor, value='Valor')
    planilha.cell(row=1, column=colEstoque, value='Estoque')
    planilha.cell(row=1, column=colLink, value='Link')
    planilha.cell(row=1, column=colData, value=data)

    ########## PERSONALIZAÇÃO #########
    #Cor de fundo dos cabeçalhos
    planilha.cell(row=1, column=colProduto).fill = azul
    planilha.cell(row=1, column=colValor).fill = azul
    planilha.cell(row=1, column=colEstoque).fill = azul
    planilha.cell(row=1, column=colLink).fill = azul
    planilha.cell(row=1, column=colData).fill = azul
    #Fonte (tamanho 11 / Branca / Negrito)
    planilha.cell(row=1, column=colProduto).font = Font(size=11, color='FFFFFF', bold=True)
    planilha.cell(row=1, column=colValor).font = Font(size=11, color='FFFFFF', bold=True)
    planilha.cell(row=1, column=colEstoque).font = Font(size=11, color='FFFFFF', bold=True)
    planilha.cell(row=1, column=colLink).font = Font(size=11, color='FFFFFF', bold=True)
    planilha.cell(row=1, column=colData).font = Font(size=11, color='FFFFFF', bold=True)
    #Alinhamento de Texto (centralizado)
    planilha.cell(row=1, column=colProduto).alignment = Alignment(horizontal='center', vertical='center')
    planilha.cell(row=1, column=colValor).alignment = Alignment(horizontal='center', vertical='center')
    planilha.cell(row=1, column=colEstoque).alignment = Alignment(horizontal='center', vertical='center')
    planilha.cell(row=1, column=colLink).alignment = Alignment(horizontal='center', vertical='center')
    planilha.cell(row=1, column=colData).alignment = Alignment(horizontal='center', vertical='center')


    #Acessa a pagina do vendedor do mercado livre
    navegador.get(linkVendedor)

    #Aguarda carregamento da pagina por completo
    navegador.implicitly_wait(2)

    #Recebe a quantidade total de anuncios
    totalAnuncios = navegador.find_element_by_class_name("ui-search-search-result__quantity-results").get_attribute("textContent")
    totalAnuncios = totalAnuncios.replace(' resultado', '')
    totalAnuncios = totalAnuncios.replace('s', '')
    totalAnuncios = int(totalAnuncios)

    #Usado para contar os itens sem perder a contagem
    itemTotal = 0 
    
    #Looping
    pagina = 1
    item = 0
    #Enquanto o anuncio for menor que 48 (48 anuncios por pagina)
    while item <= 48:
        
        #Se chegar no ultimo anuncio
        if item == 48:
            #Volta para o primeiro anuncio
            item = 0

            #Tentando mudar de pagina
            try:
                #Se estiver na primeira pagina
                if pagina == 1:
                    #Executa o script que manda para a proxima pagina
                    navegador.execute_script("document.getElementsByClassName('andes-pagination__arrow-title')[0].click()")
                #Senão
                else:
                    #Executa o script que manda para a proxima pagina
                    navegador.execute_script("document.getElementsByClassName('andes-pagination__arrow-title')[1].click()")
                
                #Aumenta o contador de paginas (indica que agora passou para proxima pagina)
                pagina += 1
                
            #Caso dê erro, pode significar que o mercado livre barrou o bot
            except:
                #Tentando mudar de pagina novamente
                try:
                    #Chama a função que burla a barreira (limpa os cookies e aguarda 80 segundos)
                    burlarBarreira(navegador)

                    #Se estiver na primeira pagina
                    if pagina == 1:
                        #Executa o script que manda para a proxima pagina
                        navegador.execute_script("document.getElementsByClassName('andes-pagination__arrow-title')[0].click()")
                    #Senão
                    else:
                        #Executa o script que manda para a proxima pagina
                        navegador.execute_script("document.getElementsByClassName('andes-pagination__arrow-title')[1].click()")
                    
                    #Aumenta o contador de paginas (indica que agora passou para proxima pagina)
                    pagina += 1

                #Se der erro, é porque nao existe mais pagina para mudar, entao sai do loop
                except:
                    #Sai do loop
                    break


        #Tenta abrir o anuncio do produto
        try:
            #Clica no produto
            navegador.execute_script("document.getElementsByClassName('ui-search-item__title ui-search-item__group__element')[" + str(item) + "].click()")
        
        #Caso dê erro, talvez seja pq o mercado livre barrou o bot
        except:
            #Tenta abrir o anuncio do produto novamente
            try:
                #Chama a função que burla a barreira ddo mercado livre (limpa os cookies e aguarda uns segundos)
                burlarBarreira(navegador)
                
                #Clica no produto
                navegador.execute_script("document.getElementsByClassName('ui-search-item__title ui-search-item__group__element')[" + str(item) + "].click()")

            #Se der erro, talvez seja pq nao tem mais produtos
            except:
                #Sai do loop
                break

        #Aguarda carregamento da pagina por completo
        navegador.implicitly_wait(2)

        #Tenta receber o nome do produto atual
        try:
            #Recebe o nome do produto
            nomeProduto = navegador.find_element_by_class_name('ui-pdp-title').text
        
        #Se der erro, significa que o mercado livre pode ter barrado o bot
        except:
            #Tenta receber o nome do produto novamente
            try:
                #Chama a função que burla a barreira (limpa os cookies e aguarda uns segundos)
                burlarBarreira(navegador)
                
                #Recebe o nome do produto
                nomeProduto = navegador.find_element_by_class_name('ui-pdp-title').text
            
            #Se der erro, exibe para o usuario
            except:
                #Exibe mensagem de erro
                pyautogui.alert(title='ERRO!', text='Algo deu errado ao receber informações do anuncio\n{}'.format(linkProduto), button='OK')
                #Fecha o navegador
                navegador.close()

        #Recebe o link do produto
        linkProduto = navegador.current_url

        #Recebe e formarta o valor do produto
        valorProduto = navegador.find_element_by_xpath('//meta[@itemprop="price"]').get_attribute("content")
        valorProduto = 'R$ ' + str(valorProduto)
        valorProduto = valorProduto.replace('.', ',')

        #Recebe o estoque disponivel do produto
        try:
            estoqueProduto = navegador.find_element_by_class_name('ui-pdp-buybox__quantity__available').text
            estoqueProduto = estoqueProduto.replace('(', '')
            estoqueProduto = estoqueProduto.replace(')', '')
        except:
            estoqueProduto = '1'

        #Remove os textos
        estoqueProduto = estoqueProduto.replace(' ', '').replace('disponíveis', '').replace('disponível', '')
        #Converte em numero
        estoqueProduto = int(estoqueProduto)

        #Recebe o numero de vendas
        try:
            vendas = navegador.find_element_by_class_name('ui-pdp-subtitle').text
        except:
            vendas = ''

        #Converte em texto
        vendas = str(vendas)
        
        #Formata o texto para deixar apenas os numeros
        vendas = vendas.replace('Novo  |  ', '')
        vendas = vendas.replace('Novo', '')
        vendas = vendas.replace(' vendidos', '')
        vendas = vendas.replace(' vendido', '')
        vendas = vendas.replace('Usado', '')
            
        #Se vendas estiver vazio, entao nao vendeu
        if vendas == '':
            vendas = '0'

        #Exibe no terminal
        print('\n\t\t\tEste é o item {} (de {}) | Página {}'.format(str(itemTotal), str(totalAnuncios), str(pagina)))
        print('\n Produto: {} \n Valor: {} \n Estoque: {} \n LINK: {} \n Vendas: {} \n'.format(nomeProduto, valorProduto, estoqueProduto, linkProduto, vendas))
        print('-' * 90)
        

        #=========================================
        #============ ANOTANDO EXCEL =============
        #=========================================
        #Anotando nome do produto
        planilha.cell(row=linha, column=colProduto, value=nomeProduto)
        #Anotando valor do produto
        planilha.cell(row=linha, column=colValor, value=valorProduto)
        #Anotando estoque
        planilha.cell(row=linha, column=colEstoque, value=estoqueProduto)
        #Anotando link do produto
        planilha.cell(row=linha, column=colLink, value=linkProduto)
        #Anotando vendas do produto
        planilha.cell(row=linha, column=colData, value=vendas)
        
        #Volta para a pagina de vendedor do mercado livre
        navegador.back()

        #Aguarda carregamento da pagina por completo
        navegador.implicitly_wait(2)
        
        #Contadores recebem mais 1
        item += 1
        itemTotal += 1
        linha += 1

    #Salvando o excel
    pastaExcel.save(caminho + '/' + nomeVendedor + '.xlsx')
    pastaExcel.close()

    #Exibe mensagem de sucesso
    pyautogui.alert(title='Fim da execução!', text='Foram registrados {} itens'.format(itemTotal), button='OK')
    print('\n\tFim da execução!\n\tForam registrados {} itens'.format(itemTotal))